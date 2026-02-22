from __future__ import annotations

import io
import logging
from decimal import Decimal
from typing import List

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from fastapi import APIRouter, Depends, File, HTTPException, UploadFile
from fastapi.responses import StreamingResponse
from sqlalchemy import delete, select
from sqlalchemy.ext.asyncio import AsyncSession

from app.core.deps import get_current_user, get_db
from app.models.service import Service
from app.models.user import User
from app.schemas.service import (
    BulkDeleteRequest,
    BulkExportRequest,
    BulkPriceChangeRequest,
    MaterialItem,
    ServiceCreate,
    ServiceResponse,
    ServiceUpdate,
)
from app.services.ai_provider import analyze_image, get_provider, parse_ai_json
from app.services.nhsu import get_tax_rates

logger = logging.getLogger(__name__)

router = APIRouter()

# ── Фінансові розрахунки ─────────────────────────────────────────────


def _calc(service: Service, ep_rate: float, vz_rate: float) -> dict:
    """Розраховує всі фінансові поля послуги."""
    price = float(service.price)
    materials = service.materials or []

    total_materials_cost = round(
        sum(float(m.get("cost", 0)) for m in materials), 2
    )
    ep_amount = round(price * ep_rate / 100, 2)
    vz_amount = round(price * vz_rate / 100, 2)
    total_costs = round(total_materials_cost + ep_amount + vz_amount, 2)
    net_income = round(price - total_costs, 2)
    doctor_income = round(net_income / 2, 2)
    org_income = round(net_income / 2, 2)

    return {
        "total_materials_cost": total_materials_cost,
        "ep_amount": ep_amount,
        "vz_amount": vz_amount,
        "total_costs": total_costs,
        "net_income": net_income,
        "doctor_income": doctor_income,
        "org_income": org_income,
    }


def _to_response(service: Service, ep_rate: float, vz_rate: float) -> ServiceResponse:
    calc = _calc(service, ep_rate, vz_rate)
    materials = [
        MaterialItem(**m) for m in (service.materials or [])
    ]
    return ServiceResponse(
        id=service.id,
        code=service.code,
        name=service.name,
        price=float(service.price),
        materials=materials,
        **calc,
    )


# ── Сидування каталогу послуг ──────────────────────────────────────────

_SEED_SERVICES = [
    {"code":"001","name":"Прийом лікаря без декларації","price":400.0,"materials":[
        {"name":"Рукавички нестерильні (1 пара)","unit":"пара","quantity":1,"cost":5.0},
        {"name":"Шпатель (1 штука)","unit":"штука","quantity":1,"cost":2.0},
        {"name":"Аскорбінова кислота (1 штука)","unit":"штука","quantity":1,"cost":12.0}]},
    {"code":"002","name":"Глюкоза крові","price":110.0,"materials":[
        {"name":"Рукавички нестерильні (1 пара)","unit":"пара","quantity":1,"cost":5.0},
        {"name":"Тест-смужка (1 штука)","unit":"штука","quantity":1,"cost":10.0},
        {"name":"Спиртова серветка (1 штука)","unit":"штука","quantity":1,"cost":1.0},
        {"name":"Ланцет (1 штука)","unit":"штука","quantity":1,"cost":1.0}]},
    {"code":"003","name":"Глюкоза крові з навантаженням","price":250.0,"materials":[
        {"name":"Рукавички нестерильні (2 пари)","unit":"пара","quantity":2,"cost":10.0},
        {"name":"Спиртова серветка (2 штуки)","unit":"штука","quantity":2,"cost":2.0},
        {"name":"Тест-смужка (2 штуки)","unit":"штука","quantity":2,"cost":20.0},
        {"name":"Ланцет (2 штуки)","unit":"штука","quantity":2,"cost":2.0},
        {"name":"Глюкоза-тест порошок для орального розчину (1 штука)","unit":"штука","quantity":1,"cost":70.0},
        {"name":"Шпатель (1 штука)","unit":"штука","quantity":1,"cost":2.0},
        {"name":"Одноразовий стаканчик (1 штука)","unit":"штука","quantity":1,"cost":1.0}]},
    {"code":"004","name":"Електрокардіографія","price":150.0,"materials":[
        {"name":"Амортизація та стрічка","unit":"послуга","quantity":1,"cost":50.0},
        {"name":"Рукавички нестерильні (1 пара)","unit":"пара","quantity":1,"cost":5.0}]},
    {"code":"005","name":"Швидкий тест на антиген COVID-19","price":250.0,"materials":[
        {"name":"Експрес-тест на антиген COVID-19 (1 штука)","unit":"штука","quantity":1,"cost":50.0},
        {"name":"Рукавички нестерильні (1 пара)","unit":"пара","quantity":1,"cost":5.0},
        {"name":"Маска медична 3-шарова (1 штука)","unit":"штука","quantity":1,"cost":1.50}]},
    {"code":"006","name":"Загальний аналіз сечі","price":150.0,"materials":[
        {"name":"Рукавички нестерильні (2 пари)","unit":"пара","quantity":2,"cost":10.0},
        {"name":"Тест-смужки U-11, Mindray (1 штука)","unit":"штука","quantity":1,"cost":20.0},
        {"name":"Амортизація","unit":"послуга","quantity":1,"cost":15.0}]},
    {"code":"007","name":"Внутрішньовенне введення лікарського засобу","price":100.0,"materials":[
        {"name":"Шприц 10,0 мл (2 штуки)","unit":"штука","quantity":2,"cost":8.0},
        {"name":"Рукавички нестерильні (1 пара)","unit":"пара","quantity":1,"cost":5.0},
        {"name":"Спиртова серветка (2 штуки)","unit":"штука","quantity":2,"cost":2.0},
        {"name":"Лейкопластир (1 штука)","unit":"штука","quantity":1,"cost":2.50},
        {"name":"Шприц 20,0 мл (1 штука)","unit":"штука","quantity":1,"cost":5.0}]},
    {"code":"008","name":"Внутрішньом\u2019язове введення лікарського засобу","price":80.0,"materials":[
        {"name":"Шприц 5,0 мл (2 штуки)","unit":"штука","quantity":2,"cost":8.0},
        {"name":"Рукавички нестерильні (1 пара)","unit":"пара","quantity":1,"cost":5.0},
        {"name":"Спиртова серветка (2 штуки)","unit":"штука","quantity":2,"cost":2.0},
        {"name":"Шприц 2,0 мл (2 штуки)","unit":"штука","quantity":2,"cost":6.0}]},
    {"code":"009","name":"Підшкірне введення лікарського засобу","price":80.0,"materials":[
        {"name":"Шприц 5,0 мл (1 штука)","unit":"штука","quantity":1,"cost":4.0},
        {"name":"Рукавички нестерильні (1 пара)","unit":"пара","quantity":1,"cost":5.0},
        {"name":"Спиртова серветка (2 штуки)","unit":"штука","quantity":2,"cost":2.0},
        {"name":"Шприц 2,0 мл (1 штука)","unit":"штука","quantity":1,"cost":3.0}]},
    {"code":"010","name":"Інфузійне введення лікарського засобу","price":200.0,"materials":[
        {"name":"Шприц 5,0 мл (1 штука)","unit":"штука","quantity":1,"cost":4.0},
        {"name":"Рукавички нестерильні (1 пара)","unit":"пара","quantity":1,"cost":5.0},
        {"name":"Спиртова серветка (2 штуки)","unit":"штука","quantity":2,"cost":2.0},
        {"name":"Шприц 10,0 мл (1 штука)","unit":"штука","quantity":1,"cost":4.0},
        {"name":"Інфузійна система (1 штука)","unit":"штука","quantity":1,"cost":15.0},
        {"name":"Пластир для фіксації канюлі внутрішньовенної (1 штука)","unit":"штука","quantity":1,"cost":4.0}]},
    {"code":"011","name":"Взяття біологічного матеріалу (венозна кров)","price":100.0,"materials":[
        {"name":"Шприц 5,0 мл (1 штука)","unit":"штука","quantity":1,"cost":4.0},
        {"name":"Рукавички нестерильні (1 пара)","unit":"пара","quantity":1,"cost":5.0},
        {"name":"Спиртова серветка (2 штуки)","unit":"штука","quantity":2,"cost":2.0},
        {"name":"Шприц інсуліновий 1,0 мл (1 штука)","unit":"штука","quantity":1,"cost":5.0}]},
    {"code":"012","name":"Прийом лікаря-спеціаліста","price":500.0,"materials":[
        {"name":"Рукавички нестерильні (1 пара)","unit":"пара","quantity":1,"cost":5.0},
        {"name":"Шпатель (1 штука)","unit":"штука","quantity":1,"cost":2.0},
        {"name":"Аскорбінова кислота (1 штука)","unit":"штука","quantity":1,"cost":12.0}]},
    {"code":"013","name":"Консультація лікаря на дому","price":600.0,"materials":[
        {"name":"Рукавички нестерильні (1 пара)","unit":"пара","quantity":1,"cost":5.0},
        {"name":"Шпатель (1 штука)","unit":"штука","quantity":1,"cost":2.0},
        {"name":"Послуги таксі — середній показник","unit":"послуга","quantity":1,"cost":150.0}]},
    {"code":"014","name":"Комбінований швидкий тест для визначення антигена COVID-19 та грипу А/В","price":350.0,"materials":[
        {"name":"Комбінований швидкий тест для визначення антигена COVID-19 та грипу А/В (1 штука)","unit":"штука","quantity":1,"cost":105.0},
        {"name":"Рукавички нестерильні (1 пара)","unit":"пара","quantity":1,"cost":5.0},
        {"name":"Маска медична 3-шарова (1 штука)","unit":"штука","quantity":1,"cost":1.50}]},
    {"code":"015","name":"Швидкий тест для діагностики стрептококової ангіни","price":400.0,"materials":[
        {"name":"Швидкий тест для діагностики стрептококової ангіни CITO TEST STREP A (1 штука)","unit":"штука","quantity":1,"cost":255.0},
        {"name":"Рукавички нестерильні (1 пара)","unit":"пара","quantity":1,"cost":5.0},
        {"name":"Маска медична 3-шарова (1 штука)","unit":"штука","quantity":1,"cost":1.50}]},
    {"code":"016","name":"Визначення рівня сечової кислоти у крові тест-смужками","price":200.0,"materials":[
        {"name":"Рукавички нестерильні (1 пара)","unit":"пара","quantity":1,"cost":5.0},
        {"name":"Тест-смужка (1 штука)","unit":"штука","quantity":1,"cost":35.0},
        {"name":"Спиртова серветка (1 штука)","unit":"штука","quantity":1,"cost":1.0},
        {"name":"Ланцет (1 штука)","unit":"штука","quantity":1,"cost":1.0}]},
    {"code":"017","name":"Швидкий тест для визначення С-реактивного білка","price":250.0,"materials":[
        {"name":"Швидкий тест для визначення С-реактивного білка (1 штука)","unit":"штука","quantity":1,"cost":75.0},
        {"name":"Рукавички нестерильні (1 пара)","unit":"пара","quantity":1,"cost":5.0},
        {"name":"Спиртова серветка (1 штука)","unit":"штука","quantity":1,"cost":1.0},
        {"name":"Ланцет (1 штука)","unit":"штука","quantity":1,"cost":1.0}]},
    {"code":"018","name":"Швидкий тест на феритин","price":350.0,"materials":[
        {"name":"Швидкий тест на феритин (1 штука)","unit":"штука","quantity":1,"cost":113.0},
        {"name":"Рукавички нестерильні (1 пара)","unit":"пара","quantity":1,"cost":5.0},
        {"name":"Спиртова серветка (1 штука)","unit":"штука","quantity":1,"cost":1.0},
        {"name":"Ланцет (1 штука)","unit":"штука","quantity":1,"cost":1.0}]},
    {"code":"019","name":"Швидкий тест на вітамін D","price":500.0,"materials":[
        {"name":"Швидкий тест на вітамін D (1 штука)","unit":"штука","quantity":1,"cost":148.0},
        {"name":"Рукавички нестерильні (1 пара)","unit":"пара","quantity":1,"cost":5.0},
        {"name":"Спиртова серветка (1 штука)","unit":"штука","quantity":1,"cost":1.0},
        {"name":"Ланцет (1 штука)","unit":"штука","quantity":1,"cost":1.0}]},
    {"code":"020","name":"Швидкий тест на виявлення простат-специфічного антигену (ПСА)","price":340.0,"materials":[
        {"name":"Швидкий тест на виявлення простат-специфічного антигену (ПСА) (1 штука)","unit":"штука","quantity":1,"cost":72.0},
        {"name":"Рукавички нестерильні (1 пара)","unit":"пара","quantity":1,"cost":5.0},
        {"name":"Спиртова серветка (1 штука)","unit":"штука","quantity":1,"cost":1.0},
        {"name":"Ланцет (1 штука)","unit":"штука","quantity":1,"cost":1.0}]},
    {"code":"021","name":"Тест швидкий комбінований для виявлення ВІЛ 1/2, гепатит С, гепатит В, сифіліс","price":450.0,"materials":[
        {"name":"Тест швидкий комбінований для виявлення ВІЛ 1/2, гепатит С, гепатит В, сифіліс (1 штука)","unit":"штука","quantity":1,"cost":80.0},
        {"name":"Рукавички нестерильні (1 пара)","unit":"пара","quantity":1,"cost":5.0},
        {"name":"Спиртова серветка (1 штука)","unit":"штука","quantity":1,"cost":1.0},
        {"name":"Ланцет (1 штука)","unit":"штука","quantity":1,"cost":1.0}]},
    {"code":"022","name":"Визначення рівня гемоглобіну в крові тест-смужками","price":100.0,"materials":[
        {"name":"Рукавички нестерильні (1 пара)","unit":"пара","quantity":1,"cost":5.0},
        {"name":"Тест-смужка (1 штука)","unit":"штука","quantity":1,"cost":39.0},
        {"name":"Спиртова серветка (1 штука)","unit":"штука","quantity":1,"cost":1.0},
        {"name":"Ланцет (1 штука)","unit":"штука","quantity":1,"cost":1.0}]},
]


@router.post("/seed", status_code=200)
async def seed_services(
    db: AsyncSession = Depends(get_db),
    user: User = Depends(get_current_user),
):
    """Видалити всі існуючі послуги та створити 22 стандартних послуги."""
    await db.execute(delete(Service).where(Service.user_id == user.id))
    created = []
    for s in _SEED_SERVICES:
        svc = Service(
            user_id=user.id,
            code=s["code"],
            name=s["name"],
            price=s["price"],
            materials=s["materials"],
        )
        db.add(svc)
        created.append(s["code"])
    await db.commit()
    return {"deleted": "all", "created": len(created), "codes": created}


# ── AI-аналіз зображень послуг ─────────────────────────────────────

_SERVICE_AI_SYSTEM = """\
Ти — спеціалізований OCR-асистент для розпізнавання прайс-листів медичних послуг.
Твоя задача — точно розпізнати дані послуг з зображення та повернути їх у форматі JSON.
Відповідай ТІЛЬКИ валідним JSON без markdown, коментарів чи пояснень."""

_SERVICE_AI_PROMPT = """\
<context>
Це зображення прайс-листа, таблиці або списку медичних послуг.
На зображенні можуть бути: назви послуг, коди послуг, ціни, матеріали та їх вартість.
</context>

<task>
Розпізнай ВСІ послуги з зображення. Для кожної послуги вилучи:

1. **code** — код або номер послуги (якщо є). Якщо коду немає — генеруй порядковий номер як рядок: "001", "002", "003"...
2. **name** — повна назва послуги українською мовою
3. **price** — ціна послуги для клієнта (в грн). Якщо ціна не вказана — постав 0.
4. **materials** — список матеріалів/витрат (якщо видно на зображенні):
   - name: назва матеріалу
   - unit: одиниця виміру (штука, пара, мл тощо)
   - quantity: кількість (число)
   - cost: загальна вартість цього матеріалу (грн, число)
   Якщо матеріали не вказані — повертай порожній масив [].
</task>

<rules>
- Розпізнай ВСІ послуги, навіть якщо їх десятки
- Ціни — числа (float), наприклад 150.0, 2500.00
- quantity та cost — числа (float)
- Якщо код не видно — використай порядкові номери: "001", "002" тощо
- Якщо назва частково видима — вкажи те, що вдалося розпізнати
- НЕ вигадуй послуги, яких немає на зображенні
- confidence: "high" = все чітко видно, "medium" = є неточності, "low" = дані погано розпізнані
</rules>

<output_format>
Поверни ТІЛЬКИ JSON:
{
  "services": [
    {
      "code": "001",
      "name": "Назва послуги",
      "price": 100.0,
      "materials": [
        {"name": "Матеріал", "unit": "штука", "quantity": 1, "cost": 5.0}
      ]
    }
  ],
  "total_found": 5,
  "confidence": "high",
  "notes": ""
}
</output_format>"""


@router.post("/analyze-image")
async def analyze_services_image(
    images: List[UploadFile] = File(...),
    db: AsyncSession = Depends(get_db),
    user: User = Depends(get_current_user),
):
    """Аналізує зображення прайс-листа через AI та повертає розпізнані послуги."""
    provider, api_key = await get_provider(db, user.id)

    if not provider:
        raise HTTPException(
            status_code=503,
            detail=(
                "AI сервіс не налаштований. "
                "Додайте хоча б один API ключ у Налаштуваннях: "
                "Anthropic (Claude), OpenAI (ChatGPT) або xAI (Grok)."
            ),
        )

    all_services = []
    notes_parts = []

    for image_file in images:
        raw = await image_file.read()
        media_type = image_file.content_type or "image/png"

        if len(raw) > 10 * 1024 * 1024:
            notes_parts.append(f"Файл {image_file.filename} перевищує 10 MB — пропущено")
            continue

        try:
            text = await analyze_image(
                provider, api_key, raw, media_type,
                _SERVICE_AI_SYSTEM, _SERVICE_AI_PROMPT,
            )
            data = parse_ai_json(text, {"services": [], "confidence": "low", "notes": "Не вдалося розпізнати"})

            services_list = data.get("services", [])
            for svc in services_list:
                # Normalise materials
                materials = []
                for m in svc.get("materials", []):
                    materials.append({
                        "name": str(m.get("name", "")),
                        "unit": str(m.get("unit", "")),
                        "quantity": float(m.get("quantity", 0)),
                        "cost": float(m.get("cost", 0)),
                    })
                all_services.append({
                    "code": str(svc.get("code", "")),
                    "name": str(svc.get("name", "")),
                    "price": float(svc.get("price", 0)),
                    "materials": materials,
                })

            if data.get("notes"):
                notes_parts.append(f"{image_file.filename}: {data['notes']}")

        except Exception as exc:
            logger.exception("AI analysis failed for %s", image_file.filename)
            notes_parts.append(f"Помилка аналізу {image_file.filename}: {exc}")

    return {
        "services": all_services,
        "total_found": len(all_services),
        "provider": provider,
        "notes": "; ".join(notes_parts) if notes_parts else "",
    }


@router.post("/bulk-create", response_model=list[ServiceResponse])
async def bulk_create_services(
    services_in: list[ServiceCreate],
    db: AsyncSession = Depends(get_db),
    user: User = Depends(get_current_user),
):
    """Масове створення послуг. Пропускає дублікати за кодом."""
    # Отримати існуючі коди
    result = await db.execute(
        select(Service.code).where(Service.user_id == user.id)
    )
    existing_codes = {row[0] for row in result.all()}

    ep_rate, vz_rate = await get_tax_rates(db, user.id)
    created = []

    for svc_in in services_in:
        if svc_in.code in existing_codes:
            continue  # skip duplicate

        service = Service(
            user_id=user.id,
            code=svc_in.code,
            name=svc_in.name,
            price=svc_in.price,
            materials=[m.model_dump() for m in svc_in.materials],
        )
        db.add(service)
        existing_codes.add(svc_in.code)
        created.append(service)

    await db.commit()
    for svc in created:
        await db.refresh(svc)

    return [_to_response(s, ep_rate, vz_rate) for s in created]


# ── Масові операції (до /{service_id}!) ───────────────────────────────


@router.post("/bulk-delete", status_code=204)
async def bulk_delete_services(
    body: BulkDeleteRequest,
    db: AsyncSession = Depends(get_db),
    user: User = Depends(get_current_user),
):
    """Масове видалення послуг."""
    if not body.ids:
        return
    result = await db.execute(
        select(Service).where(
            Service.id.in_(body.ids), Service.user_id == user.id
        )
    )
    services = result.scalars().all()
    for svc in services:
        await db.delete(svc)
    await db.commit()


@router.post("/bulk-price-change", response_model=list[ServiceResponse])
async def bulk_price_change(
    body: BulkPriceChangeRequest,
    db: AsyncSession = Depends(get_db),
    user: User = Depends(get_current_user),
):
    """Масова зміна ціни: нова = поточна × (1 + % / 100)."""
    if not body.ids:
        return []
    ep_rate, vz_rate = await get_tax_rates(db, user.id)
    result = await db.execute(
        select(Service).where(
            Service.id.in_(body.ids), Service.user_id == user.id
        )
    )
    services = result.scalars().all()
    for svc in services:
        new_price = float(svc.price) * (1 + body.percent / 100)
        svc.price = round(new_price, 2)
    await db.commit()
    for svc in services:
        await db.refresh(svc)
    return [_to_response(s, ep_rate, vz_rate) for s in services]


@router.post("/export")
async def export_services(
    body: BulkExportRequest,
    db: AsyncSession = Depends(get_db),
    user: User = Depends(get_current_user),
):
    """Експорт вибраних послуг у файл Excel (.xlsx) з усіма розрахунками та розхідниками."""
    ep_rate, vz_rate = await get_tax_rates(db, user.id)

    query = select(Service).where(Service.user_id == user.id)
    if body.ids:
        query = query.where(Service.id.in_(body.ids))
    result = await db.execute(query.order_by(Service.code))
    services = result.scalars().all()

    wb = openpyxl.Workbook()
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="2D3748", end_color="2D3748", fill_type="solid")
    thin_border = Border(
        bottom=Side(style="thin", color="E2E8F0"),
    )

    def style_header(ws, col_count):
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    def auto_width(ws):
        for col in ws.columns:
            max_len = max(len(str(cell.value or "")) for cell in col)
            ws.column_dimensions[col[0].column_letter].width = max(max_len + 3, 12)

    # ── Аркуш 1: Послуги ──
    ws1 = wb.active
    ws1.title = "Послуги"

    headers_main = [
        "№", "Код", "Назва послуги", "Ціна (грн)",
        "Витрати на матеріали (грн)", f"Єдиний податок {ep_rate}% (грн)",
        f"Військовий збір {vz_rate}% (грн)", "Сумарні витрати (грн)",
        "Чистий дохід (грн)", "Дохід лікаря (грн)", "Дохід організації (грн)",
    ]
    ws1.append(headers_main)
    style_header(ws1, len(headers_main))

    for idx, svc in enumerate(services, 1):
        c = _calc(svc, ep_rate, vz_rate)
        ws1.append([
            idx,
            svc.code,
            svc.name,
            float(svc.price),
            c["total_materials_cost"],
            c["ep_amount"],
            c["vz_amount"],
            c["total_costs"],
            c["net_income"],
            c["doctor_income"],
            c["org_income"],
        ])

    # Числовий формат для грошових стовпців (D-K)
    for row in ws1.iter_rows(min_row=2, min_col=4, max_col=11):
        for cell in row:
            cell.number_format = '#,##0.00'

    auto_width(ws1)

    # ── Аркуш 2: Розхідники ──
    ws2 = wb.create_sheet("Розхідники")

    # Визначаємо макс. кількість матеріалів серед усіх послуг
    max_materials = max((len(svc.materials or []) for svc in services), default=0)

    # Заголовок: Код | Назва | К-ть розхідників | Матеріал 1 (назва) | Од. | К-ть | Вартість | Матеріал 2 ...
    headers_mat = ["Код", "Назва послуги", "К-ть розхідників"]
    for i in range(1, max_materials + 1):
        headers_mat.extend([
            f"Матеріал {i} — назва",
            f"Матеріал {i} — одиниця",
            f"Матеріал {i} — кількість",
            f"Матеріал {i} — вартість (грн)",
        ])
    ws2.append(headers_mat)
    style_header(ws2, len(headers_mat))

    for svc in services:
        materials = svc.materials or []
        row_data = [svc.code, svc.name, len(materials)]
        for m in materials:
            row_data.extend([
                m.get("name", ""),
                m.get("unit", ""),
                m.get("quantity", 0),
                float(m.get("cost", 0)),
            ])
        ws2.append(row_data)

    # Числовий формат для стовпців вартості (кожен 4-й починаючи з 7-го: G, K, O, ...)
    for mat_idx in range(max_materials):
        cost_col = 4 + mat_idx * 4 + 3  # колонка вартості (1-indexed: 7, 11, 15, ...)
        for row in ws2.iter_rows(min_row=2, min_col=cost_col, max_col=cost_col):
            for cell in row:
                if cell.value is not None:
                    cell.number_format = '#,##0.00'

    auto_width(ws2)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=services.xlsx"},
    )


# ── CRUD ──────────────────────────────────────────────────────────────


@router.get("/", response_model=list[ServiceResponse])
async def list_services(
    db: AsyncSession = Depends(get_db),
    user: User = Depends(get_current_user),
):
    ep_rate, vz_rate = await get_tax_rates(db, user.id)
    result = await db.execute(
        select(Service)
        .where(Service.user_id == user.id)
        .order_by(Service.code)
    )
    services = result.scalars().all()
    return [_to_response(s, ep_rate, vz_rate) for s in services]


@router.post("/", response_model=ServiceResponse, status_code=201)
async def create_service(
    service_in: ServiceCreate,
    db: AsyncSession = Depends(get_db),
    user: User = Depends(get_current_user),
):
    # Перевірка унікальності коду в межах користувача
    existing = await db.execute(
        select(Service).where(
            Service.user_id == user.id, Service.code == service_in.code
        )
    )
    if existing.scalar_one_or_none():
        raise HTTPException(
            status_code=400,
            detail=f"Послуга з кодом '{service_in.code}' вже існує",
        )

    service = Service(
        user_id=user.id,
        code=service_in.code,
        name=service_in.name,
        price=service_in.price,
        materials=[m.model_dump() for m in service_in.materials],
    )
    db.add(service)
    await db.commit()
    await db.refresh(service)

    ep_rate, vz_rate = await get_tax_rates(db, user.id)
    return _to_response(service, ep_rate, vz_rate)


@router.put("/{service_id}", response_model=ServiceResponse)
async def update_service(
    service_id: int,
    service_in: ServiceUpdate,
    db: AsyncSession = Depends(get_db),
    user: User = Depends(get_current_user),
):
    result = await db.execute(
        select(Service).where(
            Service.id == service_id, Service.user_id == user.id
        )
    )
    service = result.scalar_one_or_none()
    if not service:
        raise HTTPException(status_code=404, detail="Послугу не знайдено")

    data = service_in.model_dump(exclude_unset=True)

    # Перевірка унікальності нового коду
    if "code" in data and data["code"] != service.code:
        existing = await db.execute(
            select(Service).where(
                Service.user_id == user.id,
                Service.code == data["code"],
                Service.id != service_id,
            )
        )
        if existing.scalar_one_or_none():
            raise HTTPException(
                status_code=400,
                detail=f"Послуга з кодом '{data['code']}' вже існує",
            )

    if "materials" in data:
        # Конвертуємо MaterialItem → dict
        data["materials"] = [
            m.model_dump() if isinstance(m, MaterialItem) else m
            for m in data["materials"]
        ]

    for field, value in data.items():
        setattr(service, field, value)

    await db.commit()
    await db.refresh(service)

    ep_rate, vz_rate = await get_tax_rates(db, user.id)
    return _to_response(service, ep_rate, vz_rate)


@router.delete("/{service_id}", status_code=204)
async def delete_service(
    service_id: int,
    db: AsyncSession = Depends(get_db),
    user: User = Depends(get_current_user),
):
    result = await db.execute(
        select(Service).where(
            Service.id == service_id, Service.user_id == user.id
        )
    )
    service = result.scalar_one_or_none()
    if not service:
        raise HTTPException(status_code=404, detail="Послугу не знайдено")
    await db.delete(service)
    await db.commit()
