"""Модуль щомісячного обліку платних послуг.

Захищені маршрути  : /api/monthly-services/*
Публічний маршрут  : GET /api/monthly-services/share/{token}/view  (без авторизації)
"""
from __future__ import annotations

import io
import secrets
import time
from collections import defaultdict
from datetime import datetime, timedelta, timezone
from typing import Optional

import openpyxl
from fastapi import APIRouter, Depends, File, Form, HTTPException, Query, UploadFile
from fastapi.responses import StreamingResponse
from pydantic import BaseModel, field_validator
from sqlalchemy import delete, func, select
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.orm.attributes import flag_modified

from app.core.deps import get_current_user, get_db
from app.models.doctor import Doctor
from app.models.monthly_service import (
    MonthlyPaidServiceEntry,
    MonthlyPaidServicesReport,
    MonthlyPeriodCash,
    MonthlyServiceLock,
)
from app.models.service import Service
from app.models.share_report import ShareReport
from app.models.user import User
from app.services.ai_provider import analyze_image as ai_analyze_image, get_provider, parse_ai_json
from app.services.nhsu import get_tax_rates
from app.schemas.monthly_service import (
    AnalyticsResponse,
    DashboardData,
    DoctorBreakdown,
    EntryResponse,
    ExportRequest,
    MonthlyTrendRow,
    PeriodInfoResponse,
    PublicShareData,
    ReportCreate,
    ReportResponse,
    ReportUpdate,
    ServiceTableRow,
    ShareCreate,
    ShareResponse,
    TopMaterialRow,
)

router = APIRouter()

MONTHS_UA = [
    "", "Січень", "Лютий", "Березень", "Квітень",
    "Травень", "Червень", "Липень", "Серпень",
    "Вересень", "Жовтень", "Листопад", "Грудень",
]

# ── Утиліти ─────────────────────────────────────────────────────────


async def _get_doctor_name(db: AsyncSession, doctor_id: int) -> str:
    r = await db.execute(select(Doctor).where(Doctor.id == doctor_id))
    d = r.scalar_one_or_none()
    return d.full_name if d else "—"


def _svc_materials_cost(svc: Service) -> float:
    return sum(
        float(m.get("quantity", 0)) * float(m.get("cost", 0))
        for m in (svc.materials or [])
    )


def _calc_row(price: float, mat_cost: float, qty: int, ep_r: float, vz_r: float) -> dict:
    P, M, Q = price, mat_cost, qty
    EP = round(P * ep_r / 100, 2)
    VZ = round(P * vz_r / 100, 2)
    to_split1 = round(P - M - EP - VZ, 2)
    ts = round(to_split1 * Q, 2)
    dr = round(ts / 2, 2)
    org = round(ts - dr, 2)
    return {
        "sum": round(P * Q, 2),
        "materials": round(M * Q, 2),
        "ep_amount": round(EP * Q, 2),
        "vz_amount": round(VZ * Q, 2),
        "total_costs": round((M + EP + VZ) * Q, 2),
        "to_split": ts,
        "doctor_income": dr,
        "org_income": org,
    }


async def _get_period_cash(
    db: AsyncSession, user_id: int, year: int, month: int
) -> Optional[float]:
    """Повертає готівку за місяць або None якщо ще не внесена."""
    r = await db.execute(
        select(MonthlyPeriodCash).where(
            MonthlyPeriodCash.user_id == user_id,
            MonthlyPeriodCash.period_year == year,
            MonthlyPeriodCash.period_month == month,
        )
    )
    record = r.scalar_one_or_none()
    return float(record.amount) if record is not None else None


async def _set_period_cash(
    db: AsyncSession, user_id: int, year: int, month: int, amount: float
) -> tuple[float, bool]:
    """Встановлює готівку за місяць. Повертає (amount, created).

    Якщо запис вже існує — повертає існуюче значення без змін (False).
    """
    r = await db.execute(
        select(MonthlyPeriodCash).where(
            MonthlyPeriodCash.user_id == user_id,
            MonthlyPeriodCash.period_year == year,
            MonthlyPeriodCash.period_month == month,
        )
    )
    existing = r.scalar_one_or_none()
    if existing is not None:
        return float(existing.amount), False   # вже є — не перезаписуємо

    cash_rec = MonthlyPeriodCash(
        user_id=user_id,
        period_year=year,
        period_month=month,
        amount=amount,
    )
    db.add(cash_rec)
    return amount, True


class ServiceLockRequest(BaseModel):
    year: int
    month: int

    @field_validator("month")
    @classmethod
    def check_month(cls, v: int) -> int:
        if v < 1 or v > 12:
            raise ValueError("Місяць повинен бути від 1 до 12")
        return v


async def _check_service_period_lock(db: AsyncSession, user_id: int, year: int, month: int) -> None:
    """Перевіряє чи заблокований період платних послуг. Кидає 423 якщо так."""
    res = await db.execute(
        select(MonthlyServiceLock).where(
            MonthlyServiceLock.user_id == user_id,
            MonthlyServiceLock.year == year,
            MonthlyServiceLock.month == month,
        )
    )
    if res.scalar_one_or_none() is not None:
        raise HTTPException(
            status_code=423,
            detail=f"Період {month:02d}/{year} заблоковано. Розблокуйте для редагування.",
        )


async def _build_analytics(
    db: AsyncSession,
    user_id: int,
    year: int,
    month: int,
    doctor_id: Optional[int],
) -> AnalyticsResponse:
    ep_rate, vz_rate = await get_tax_rates(db, user_id)

    # ── Поточний місяць: звіти ──
    rq = select(MonthlyPaidServicesReport).where(
        MonthlyPaidServicesReport.user_id == user_id,
        MonthlyPaidServicesReport.year == year,
        MonthlyPaidServicesReport.month == month,
    )
    if doctor_id:
        rq = rq.where(MonthlyPaidServicesReport.doctor_id == doctor_id)
    cur_reports_r = await db.execute(rq)
    cur_reports = cur_reports_r.scalars().all()

    report_ids = [r.id for r in cur_reports]

    # ── Записи послуг ──
    entries_r = await db.execute(
        select(MonthlyPaidServiceEntry).where(
            MonthlyPaidServiceEntry.report_id.in_(report_ids)
        )
        if report_ids
        else select(MonthlyPaidServiceEntry).where(False)
    )
    entries = entries_r.scalars().all()

    # ── Довідник послуг (всі поточного user) ──
    svcs_r = await db.execute(
        select(Service).where(Service.user_id == user_id)
    )
    svcs_all = {s.id: s for s in svcs_r.scalars().all()}

    # Лікарі звітів
    report_by_id = {r.id: r for r in cur_reports}
    doctor_ids = list({r.doctor_id for r in cur_reports})
    doctors_r = await db.execute(
        select(Doctor).where(Doctor.id.in_(doctor_ids)) if doctor_ids else select(Doctor).where(False)
    )
    doctors_map: dict[int, str] = {d.id: d.full_name for d in doctors_r.scalars().all()}

    # ── Агрегація по послугах ──
    svc_qty: dict[int, int] = defaultdict(int)                  # service_id → qty
    svc_by_doc: dict[int, list[tuple[int, int]]] = defaultdict(list)  # service_id → [(doc_id, qty)]

    for e in entries:
        rep = report_by_id.get(e.report_id)
        if rep:
            svc_qty[e.service_id] += e.quantity
            svc_by_doc[e.service_id].append((rep.doctor_id, e.quantity))

    # ── Підсумки (per-entry для точного округлення) ──
    _CALC_KEYS = ("sum", "materials", "ep_amount", "vz_amount", "total_costs",
                  "to_split", "doctor_income", "org_income")
    totals = dict(sum=0.0, materials=0.0, ep=0.0, vz=0.0, to_split=0.0, dr=0.0, org=0.0)
    services_table: list[ServiceTableRow] = []
    mat_agg: dict[str, dict] = {}   # material name → {unit, qty, cost}

    for svc_id, qty in svc_qty.items():
        if qty == 0:
            continue
        svc = svcs_all.get(svc_id)
        if not svc:
            continue

        M = _svc_materials_cost(svc)

        # Суммуємо per-entry результати для кожної послуги
        s_agg = {k: 0.0 for k in _CALC_KEYS}
        for _doc_id, entry_qty in svc_by_doc[svc_id]:
            c = _calc_row(float(svc.price), M, entry_qty, ep_rate, vz_rate)
            for k in _CALC_KEYS:
                s_agg[k] += c[k]
        for k in _CALC_KEYS:
            s_agg[k] = round(s_agg[k], 2)

        totals["sum"] += s_agg["sum"]
        totals["materials"] += s_agg["materials"]
        totals["ep"] += s_agg["ep_amount"]
        totals["vz"] += s_agg["vz_amount"]
        totals["to_split"] += s_agg["to_split"]
        totals["dr"] += s_agg["doctor_income"]
        totals["org"] += s_agg["org_income"]

        by_doc = [
            DoctorBreakdown(doctor_id=did, doctor_name=doctors_map.get(did, "—"), quantity=q)
            for did, q in svc_by_doc[svc_id]
        ]
        services_table.append(ServiceTableRow(
            service_id=svc_id,
            code=svc.code,
            name=svc.name,
            price=float(svc.price),
            total_quantity=qty,
            by_doctor=by_doc,
            **s_agg,
        ))

        # Агрегація матеріалів для ТОП
        for mat in (svc.materials or []):
            key = mat.get("name", "—")
            if key not in mat_agg:
                mat_agg[key] = {"unit": mat.get("unit", ""), "qty": 0.0, "cost": 0.0}
            mat_agg[key]["qty"] += float(mat.get("quantity", 0)) * qty
            mat_agg[key]["cost"] += float(mat.get("quantity", 0)) * float(mat.get("cost", 0)) * qty

    # Готівка: бере з окремої таблиці (один запис на período)
    period_cash = await _get_period_cash(db, user_id, year, month)
    cash = round(period_cash, 2) if period_cash is not None else 0.0

    # ── Попередній місяць (MoM) ──
    prev_year, prev_month = (year, month - 1) if month > 1 else (year - 1, 12)
    prev_rq = select(MonthlyPaidServicesReport).where(
        MonthlyPaidServicesReport.user_id == user_id,
        MonthlyPaidServicesReport.year == prev_year,
        MonthlyPaidServicesReport.month == prev_month,
    )
    if doctor_id:
        prev_rq = prev_rq.where(MonthlyPaidServicesReport.doctor_id == doctor_id)
    prev_reports_r = await db.execute(prev_rq)
    prev_reports = prev_reports_r.scalars().all()

    prev_revenue = 0.0
    prev_qty = 0
    if prev_reports:
        prev_entries_r = await db.execute(
            select(MonthlyPaidServiceEntry).where(
                MonthlyPaidServiceEntry.report_id.in_([r.id for r in prev_reports])
            )
        )
        for e in prev_entries_r.scalars().all():
            svc = svcs_all.get(e.service_id)
            if svc:
                prev_revenue += float(svc.price) * e.quantity
                prev_qty += e.quantity

    # ── Тренд 12 місяців (batch: 2 queries instead of 24) ──
    # Pre-compute the 12 (year, month) pairs
    trend_periods = []
    for i in range(11, -1, -1):
        m_total = month - i
        y_off = year + (m_total - 1) // 12
        m_off = ((m_total - 1) % 12) + 1
        trend_periods.append((y_off, m_off))

    # Load all trend reports in one query using OR conditions
    from sqlalchemy import or_, and_, tuple_
    trend_filter = or_(
        *[
            and_(
                MonthlyPaidServicesReport.year == yy,
                MonthlyPaidServicesReport.month == mm,
            )
            for yy, mm in trend_periods
        ]
    )
    trend_rq = select(MonthlyPaidServicesReport).where(
        MonthlyPaidServicesReport.user_id == user_id,
        trend_filter,
    )
    if doctor_id:
        trend_rq = trend_rq.where(MonthlyPaidServicesReport.doctor_id == doctor_id)
    trend_reports_r = await db.execute(trend_rq)
    all_trend_reports = trend_reports_r.scalars().all()

    # Group reports by (year, month)
    trend_reports_by_period: dict[tuple, list] = defaultdict(list)
    all_trend_report_ids = []
    for r in all_trend_reports:
        trend_reports_by_period[(r.year, r.month)].append(r)
        all_trend_report_ids.append(r.id)

    # Load all entries for all trend reports in one query
    trend_entries_by_report: dict[int, list] = defaultdict(list)
    if all_trend_report_ids:
        trend_entries_r = await db.execute(
            select(MonthlyPaidServiceEntry).where(
                MonthlyPaidServiceEntry.report_id.in_(all_trend_report_ids)
            )
        )
        for e in trend_entries_r.scalars().all():
            trend_entries_by_report[e.report_id].append(e)

    trend_rows: list[MonthlyTrendRow] = []
    for y_off, m_off in trend_periods:
        t_sum = t_mat = t_ep = t_vz = t_split = t_dr = 0.0
        t_qty = 0
        for r in trend_reports_by_period.get((y_off, m_off), []):
            for e in trend_entries_by_report.get(r.id, []):
                svc = svcs_all.get(e.service_id)
                if svc:
                    M = _svc_materials_cost(svc)
                    c = _calc_row(float(svc.price), M, e.quantity, ep_rate, vz_rate)
                    t_sum += c["sum"]
                    t_mat += c["materials"]
                    t_ep += c["ep_amount"]
                    t_vz += c["vz_amount"]
                    t_split += c["to_split"]
                    t_dr += c["doctor_income"]
                    t_qty += e.quantity
        trend_rows.append(MonthlyTrendRow(
            year=y_off, month=m_off,
            quantity=t_qty,
            sum=round(t_sum, 2), materials=round(t_mat, 2),
            ep_amount=round(t_ep, 2), vz_amount=round(t_vz, 2),
            to_split=round(t_split, 2), doctor_income=round(t_dr, 2),
        ))

    # ── ТОП матеріалів ──
    total_mat_cost = totals["materials"]
    sorted_mats = sorted(mat_agg.items(), key=lambda x: x[1]["cost"], reverse=True)
    top_materials = [
        TopMaterialRow(
            name=name,
            unit=v["unit"],
            total_quantity=round(v["qty"], 3),
            total_cost=round(v["cost"], 2),
            share_pct=round(v["cost"] / total_mat_cost * 100, 1) if total_mat_cost else 0.0,
        )
        for name, v in sorted_mats[:10]
    ]

    # ── Звіти для відповіді ──
    reports_out = await _build_report_responses(db, cur_reports, svcs_all)

    dashboard = DashboardData(
        total_revenue=round(totals["sum"], 2),
        total_quantity=sum(svc_qty.values()),
        avg_check=round(totals["sum"] / sum(svc_qty.values()), 2)
        if sum(svc_qty.values()) else 0.0,
        prev_month_revenue=round(prev_revenue, 2),
        prev_month_quantity=prev_qty,
        doctor_income=round(totals["dr"], 2),
        materials_cost=round(totals["materials"], 2),
        ep_amount=round(totals["ep"], 2),
        vz_amount=round(totals["vz"], 2),
        total_costs=round(totals["materials"] + totals["ep"] + totals["vz"], 2),
        org_income=round(totals["org"], 2),
        cash_in_register=cash,
        bank_amount=round(totals["sum"] - cash, 2),
    )

    return AnalyticsResponse(
        dashboard=dashboard,
        services_table=services_table,
        monthly_trend=trend_rows,
        top_materials=top_materials,
        reports=reports_out,
        ep_rate=ep_rate,
        vz_rate=vz_rate,
    )


async def _build_report_responses(
    db: AsyncSession,
    reports: list,
    svcs_all: dict[int, Service],
) -> list[ReportResponse]:
    if not reports:
        return []

    # Batch-load all doctors at once instead of N+1 queries
    doctor_ids = list({r.doctor_id for r in reports})
    doc_r = await db.execute(
        select(Doctor).where(Doctor.id.in_(doctor_ids))
    )
    doctors_map = {d.id: d.full_name for d in doc_r.scalars().all()}

    # Batch-load all entries for all reports at once
    report_ids = [r.id for r in reports]
    all_entries_r = await db.execute(
        select(MonthlyPaidServiceEntry).where(
            MonthlyPaidServiceEntry.report_id.in_(report_ids)
        )
    )
    entries_by_report: dict[int, list] = defaultdict(list)
    for e in all_entries_r.scalars().all():
        entries_by_report[e.report_id].append(e)

    out = []
    for rep in reports:
        entries = [
            EntryResponse(
                service_id=e.service_id,
                service_code=svcs_all[e.service_id].code if e.service_id in svcs_all else "—",
                service_name=svcs_all[e.service_id].name if e.service_id in svcs_all else "—",
                quantity=e.quantity,
            )
            for e in entries_by_report.get(rep.id, [])
        ]
        out.append(ReportResponse(
            id=rep.id,
            doctor_id=rep.doctor_id,
            doctor_name=doctors_map.get(rep.doctor_id, "—"),
            year=rep.year,
            month=rep.month,
            cash_in_register=float(rep.cash_in_register),
            status=rep.status,
            entries=entries,
            created_at=rep.created_at,
            updated_at=rep.updated_at,
        ))
    return out


# ── Ендпоінти CRUD ───────────────────────────────────────────────────


@router.get("/reports", response_model=list[ReportResponse])
async def list_reports(
    year: int = Query(...),
    month: int = Query(...),
    doctor_id: Optional[int] = Query(None),
    db: AsyncSession = Depends(get_db),
    user: User = Depends(get_current_user),
):
    rq = select(MonthlyPaidServicesReport).where(
        MonthlyPaidServicesReport.user_id == user.id,
        MonthlyPaidServicesReport.year == year,
        MonthlyPaidServicesReport.month == month,
    )
    if doctor_id:
        rq = rq.where(MonthlyPaidServicesReport.doctor_id == doctor_id)
    r = await db.execute(rq)
    reports = r.scalars().all()

    svcs_r = await db.execute(select(Service).where(Service.user_id == user.id))
    svcs_all = {s.id: s for s in svcs_r.scalars().all()}
    return await _build_report_responses(db, reports, svcs_all)


@router.post("/reports", response_model=ReportResponse, status_code=201)
async def create_report(
    body: ReportCreate,
    db: AsyncSession = Depends(get_db),
    user: User = Depends(get_current_user),
):
    await _check_service_period_lock(db, user.id, body.year, body.month)

    # Перевірка унікальності
    ex = await db.execute(
        select(MonthlyPaidServicesReport).where(
            MonthlyPaidServicesReport.user_id == user.id,
            MonthlyPaidServicesReport.doctor_id == body.doctor_id,
            MonthlyPaidServicesReport.year == body.year,
            MonthlyPaidServicesReport.month == body.month,
        )
    )
    if ex.scalar_one_or_none():
        raise HTTPException(
            400,
            detail=f"Звіт для цього лікаря та місяця вже існує",
        )

    # Зберігаємо готівку за місяць (якщо передана) — один запис на période
    if body.cash_in_register is not None:
        existing_cash = await _get_period_cash(db, user.id, body.year, body.month)
        if existing_cash is not None:
            raise HTTPException(
                409,
                detail=f"Готівка за цей місяць вже внесена ({existing_cash} грн). "
                       "Щоб змінити — відредагуйте або видаліть існуючий запис.",
            )
        await _set_period_cash(db, user.id, body.year, body.month, body.cash_in_register)

    report = MonthlyPaidServicesReport(
        user_id=user.id,
        doctor_id=body.doctor_id,
        year=body.year,
        month=body.month,
        cash_in_register=0,   # готівка зберігається в monthly_period_cash
        status="draft",
    )
    db.add(report)
    await db.flush()

    svcs_r = await db.execute(select(Service).where(Service.user_id == user.id))
    svcs_all = {s.id: s for s in svcs_r.scalars().all()}

    for e in body.entries:
        if e.quantity < 0:
            raise HTTPException(400, detail="Кількість не може бути від'ємною")
        db.add(MonthlyPaidServiceEntry(
            report_id=report.id, service_id=e.service_id, quantity=e.quantity
        ))

    await db.commit()
    await db.refresh(report)
    result = await _build_report_responses(db, [report], svcs_all)
    return result[0]


@router.get("/reports/{report_id}", response_model=ReportResponse)
async def get_report(
    report_id: int,
    db: AsyncSession = Depends(get_db),
    user: User = Depends(get_current_user),
):
    r = await db.execute(
        select(MonthlyPaidServicesReport).where(
            MonthlyPaidServicesReport.id == report_id,
            MonthlyPaidServicesReport.user_id == user.id,
        )
    )
    report = r.scalar_one_or_none()
    if not report:
        raise HTTPException(404, detail="Звіт не знайдено")
    svcs_r = await db.execute(select(Service).where(Service.user_id == user.id))
    svcs_all = {s.id: s for s in svcs_r.scalars().all()}
    result = await _build_report_responses(db, [report], svcs_all)
    return result[0]


@router.put("/reports/{report_id}", response_model=ReportResponse)
async def update_report(
    report_id: int,
    body: ReportUpdate,
    db: AsyncSession = Depends(get_db),
    user: User = Depends(get_current_user),
):
    r = await db.execute(
        select(MonthlyPaidServicesReport).where(
            MonthlyPaidServicesReport.id == report_id,
            MonthlyPaidServicesReport.user_id == user.id,
        )
    )
    report = r.scalar_one_or_none()
    if not report:
        raise HTTPException(404, detail="Звіт не знайдено")
    await _check_service_period_lock(db, user.id, report.year, report.month)
    if report.status == "final":
        raise HTTPException(403, detail="Фінальний звіт не можна редагувати")

    # Оновлюємо готівку за місяць (upsert у monthly_period_cash)
    if body.cash_in_register is not None:
        pc_r = await db.execute(
            select(MonthlyPeriodCash).where(
                MonthlyPeriodCash.user_id == user.id,
                MonthlyPeriodCash.period_year == report.year,
                MonthlyPeriodCash.period_month == report.month,
            )
        )
        pc = pc_r.scalar_one_or_none()
        if pc is not None:
            pc.amount = body.cash_in_register
        else:
            db.add(MonthlyPeriodCash(
                user_id=user.id,
                period_year=report.year,
                period_month=report.month,
                amount=body.cash_in_register,
            ))

    if body.entries is not None:
        # Видалити старі записи
        old_r = await db.execute(
            select(MonthlyPaidServiceEntry).where(
                MonthlyPaidServiceEntry.report_id == report_id
            )
        )
        for old in old_r.scalars().all():
            await db.delete(old)

        for e in body.entries:
            if e.quantity < 0:
                raise HTTPException(400, detail="Кількість не може бути від'ємною")
            db.add(MonthlyPaidServiceEntry(
                report_id=report_id, service_id=e.service_id, quantity=e.quantity
            ))

    report.updated_at = datetime.now(timezone.utc)
    await db.commit()
    await db.refresh(report)

    svcs_r = await db.execute(select(Service).where(Service.user_id == user.id))
    svcs_all = {s.id: s for s in svcs_r.scalars().all()}
    result = await _build_report_responses(db, [report], svcs_all)
    return result[0]


@router.post("/reports/{report_id}/finalize", response_model=ReportResponse)
async def finalize_report(
    report_id: int,
    db: AsyncSession = Depends(get_db),
    user: User = Depends(get_current_user),
):
    r = await db.execute(
        select(MonthlyPaidServicesReport).where(
            MonthlyPaidServicesReport.id == report_id,
            MonthlyPaidServicesReport.user_id == user.id,
        )
    )
    report = r.scalar_one_or_none()
    if not report:
        raise HTTPException(404, detail="Звіт не знайдено")
    await _check_service_period_lock(db, user.id, report.year, report.month)
    report.status = "final"
    report.updated_at = datetime.now(timezone.utc)
    await db.commit()
    await db.refresh(report)
    svcs_r = await db.execute(select(Service).where(Service.user_id == user.id))
    svcs_all = {s.id: s for s in svcs_r.scalars().all()}
    result = await _build_report_responses(db, [report], svcs_all)
    return result[0]


@router.post("/reports/{report_id}/unfinalize", response_model=ReportResponse)
async def unfinalize_report(
    report_id: int,
    db: AsyncSession = Depends(get_db),
    user: User = Depends(get_current_user),
):
    r = await db.execute(
        select(MonthlyPaidServicesReport).where(
            MonthlyPaidServicesReport.id == report_id,
            MonthlyPaidServicesReport.user_id == user.id,
        )
    )
    report = r.scalar_one_or_none()
    if not report:
        raise HTTPException(404, detail="Звіт не знайдено")
    await _check_service_period_lock(db, user.id, report.year, report.month)
    report.status = "draft"
    report.updated_at = datetime.now(timezone.utc)
    await db.commit()
    await db.refresh(report)
    svcs_r = await db.execute(select(Service).where(Service.user_id == user.id))
    svcs_all = {s.id: s for s in svcs_r.scalars().all()}
    result = await _build_report_responses(db, [report], svcs_all)
    return result[0]


@router.delete("/reports/{report_id}", status_code=204)
async def delete_report(
    report_id: int,
    db: AsyncSession = Depends(get_db),
    user: User = Depends(get_current_user),
):
    r = await db.execute(
        select(MonthlyPaidServicesReport).where(
            MonthlyPaidServicesReport.id == report_id,
            MonthlyPaidServicesReport.user_id == user.id,
        )
    )
    report = r.scalar_one_or_none()
    if not report:
        raise HTTPException(404, detail="Звіт не знайдено")
    await _check_service_period_lock(db, user.id, report.year, report.month)
    await db.delete(report)
    await db.commit()


@router.delete("/period", status_code=204)
async def delete_period_data(
    year: int = Query(...),
    month: int = Query(...),
    db: AsyncSession = Depends(get_db),
    user: User = Depends(get_current_user),
):
    """Видалити всі дані платних послуг за обраний місяць (звіти, записи та готівку)."""
    await _check_service_period_lock(db, user.id, year, month)
    # Знайти всі звіти за період
    r = await db.execute(
        select(MonthlyPaidServicesReport.id).where(
            MonthlyPaidServicesReport.user_id == user.id,
            MonthlyPaidServicesReport.year == year,
            MonthlyPaidServicesReport.month == month,
        )
    )
    report_ids = [row[0] for row in r.all()]

    if report_ids:
        # Записи видаляться каскадно (ondelete="CASCADE"), але на всяк випадок:
        await db.execute(
            delete(MonthlyPaidServiceEntry).where(
                MonthlyPaidServiceEntry.report_id.in_(report_ids)
            )
        )
        await db.execute(
            delete(MonthlyPaidServicesReport).where(
                MonthlyPaidServicesReport.id.in_(report_ids)
            )
        )

    # Видалити готівку за період
    await db.execute(
        delete(MonthlyPeriodCash).where(
            MonthlyPeriodCash.user_id == user.id,
            MonthlyPeriodCash.period_year == year,
            MonthlyPeriodCash.period_month == month,
        )
    )
    await db.commit()


@router.post("/reports/copy-previous", response_model=ReportResponse, status_code=201)
async def copy_previous_month(
    doctor_id: int = Query(...),
    year: int = Query(...),
    month: int = Query(...),
    db: AsyncSession = Depends(get_db),
    user: User = Depends(get_current_user),
):
    """Копіює записи послуг з попереднього місяця для вказаного лікаря."""
    await _check_service_period_lock(db, user.id, year, month)
    ex = await db.execute(
        select(MonthlyPaidServicesReport).where(
            MonthlyPaidServicesReport.user_id == user.id,
            MonthlyPaidServicesReport.doctor_id == doctor_id,
            MonthlyPaidServicesReport.year == year,
            MonthlyPaidServicesReport.month == month,
        )
    )
    if ex.scalar_one_or_none():
        raise HTTPException(400, detail="Звіт для цього лікаря за цей місяць вже існує")

    prev_year, prev_month = (year, month - 1) if month > 1 else (year - 1, 12)
    prev_r = await db.execute(
        select(MonthlyPaidServicesReport).where(
            MonthlyPaidServicesReport.user_id == user.id,
            MonthlyPaidServicesReport.doctor_id == doctor_id,
            MonthlyPaidServicesReport.year == prev_year,
            MonthlyPaidServicesReport.month == prev_month,
        )
    )
    prev_report = prev_r.scalar_one_or_none()
    if not prev_report:
        raise HTTPException(404, detail="Звіт за попередній місяць не знайдено")

    entries_r = await db.execute(
        select(MonthlyPaidServiceEntry).where(
            MonthlyPaidServiceEntry.report_id == prev_report.id
        )
    )
    prev_entries = entries_r.scalars().all()

    new_report = MonthlyPaidServicesReport(
        user_id=user.id,
        doctor_id=doctor_id,
        year=year,
        month=month,
        cash_in_register=0,
        status="draft",
    )
    db.add(new_report)
    await db.flush()

    for e in prev_entries:
        db.add(MonthlyPaidServiceEntry(
            report_id=new_report.id,
            service_id=e.service_id,
            quantity=e.quantity,
        ))

    await db.commit()
    await db.refresh(new_report)

    svcs_r = await db.execute(select(Service).where(Service.user_id == user.id))
    svcs_all = {s.id: s for s in svcs_r.scalars().all()}
    result = await _build_report_responses(db, [new_report], svcs_all)
    return result[0]


# ── Інформація про готівку за місяць ──────────────────────────────


@router.get("/period-info", response_model=PeriodInfoResponse)
async def get_period_info(
    year: int = Query(...),
    month: int = Query(...),
    db: AsyncSession = Depends(get_db),
    user: User = Depends(get_current_user),
):
    """Повертає інформацію про готівку за місяць.

    cash_for_period = None → каса ще не внесена, поле показувати для будь-якого лікаря.
    cash_for_period = float → каса вже внесена, поле приховати.
    """
    cash_for_period = await _get_period_cash(db, user.id, year, month)

    reps_r = await db.execute(
        select(MonthlyPaidServicesReport).where(
            MonthlyPaidServicesReport.user_id == user.id,
            MonthlyPaidServicesReport.year == year,
            MonthlyPaidServicesReport.month == month,
        )
    )
    submitted_doctor_ids = [r.doctor_id for r in reps_r.scalars().all()]

    return PeriodInfoResponse(
        last_active_doctor_id=None,   # більше не використовується
        cash_for_period=cash_for_period,
        submitted_doctor_ids=submitted_doctor_ids,
    )


# ── Аналітика ─────────────────────────────────────────────────────


# Simple TTL cache for analytics (30 seconds per user+period+doctor)
_analytics_cache: dict[str, tuple[float, object]] = {}
_ANALYTICS_CACHE_TTL = 30


@router.get("/analytics", response_model=AnalyticsResponse)
async def get_analytics(
    year: int = Query(...),
    month: int = Query(...),
    doctor_id: Optional[int] = Query(None),
    db: AsyncSession = Depends(get_db),
    user: User = Depends(get_current_user),
):
    # Перевірка фіксації: якщо є знімок — повертаємо його замість динамічного розрахунку
    lock_res = await db.execute(
        select(MonthlyServiceLock).where(
            MonthlyServiceLock.user_id == user.id,
            MonthlyServiceLock.year == year,
            MonthlyServiceLock.month == month,
        )
    )
    lock = lock_res.scalar_one_or_none()

    if lock is not None and lock.snapshot is not None:
        snapshot_data = dict(lock.snapshot)
        snapshot_data["is_locked"] = True
        return AnalyticsResponse(**snapshot_data)

    cache_key = f"{user.id}:{year}:{month}:{doctor_id}"
    now = time.time()
    if cache_key in _analytics_cache:
        cached_ts, cached_data = _analytics_cache[cache_key]
        if now - cached_ts < _ANALYTICS_CACHE_TTL:
            return cached_data

    result = await _build_analytics(db, user.id, year, month, doctor_id)

    if lock is not None:
        result.is_locked = True

    _analytics_cache[cache_key] = (now, result)
    stale = [k for k, (ts, _) in _analytics_cache.items() if now - ts > _ANALYTICS_CACHE_TTL * 2]
    for k in stale:
        _analytics_cache.pop(k, None)

    return result


# ── Фіксація / розблокування періоду ─────────────────────────────────


@router.post("/lock", status_code=200)
async def lock_period(
    body: ServiceLockRequest,
    db: AsyncSession = Depends(get_db),
    user: User = Depends(get_current_user),
):
    """Фіксує (блокує) місяць платних послуг від редагування та зберігає знімок даних."""
    existing_res = await db.execute(
        select(MonthlyServiceLock).where(
            MonthlyServiceLock.user_id == user.id,
            MonthlyServiceLock.year == body.year,
            MonthlyServiceLock.month == body.month,
        )
    )
    existing = existing_res.scalar_one_or_none()

    # Обчислюємо актуальний знімок аналітики
    analytics = await _build_analytics(db, user.id, body.year, body.month, doctor_id=None)
    analytics.is_locked = True
    analytics_dict = analytics.model_dump(mode="json")

    if existing:
        # Повторна фіксація — оновлюємо знімок
        existing.snapshot = analytics_dict
        existing.locked_at = datetime.now(timezone.utc)
        flag_modified(existing, "snapshot")
    else:
        lock = MonthlyServiceLock(
            user_id=user.id,
            year=body.year,
            month=body.month,
            snapshot=analytics_dict,
        )
        db.add(lock)

    await db.commit()

    # Інвалідуємо кеш аналітики для цього періоду
    for key in list(_analytics_cache.keys()):
        if key.startswith(f"{user.id}:{body.year}:{body.month}:"):
            _analytics_cache.pop(key, None)

    return {"is_locked": True}


@router.delete("/lock", status_code=200)
async def unlock_period(
    year: int = Query(...),
    month: int = Query(...),
    db: AsyncSession = Depends(get_db),
    user: User = Depends(get_current_user),
):
    """Розблоковує місяць платних послуг для редагування."""
    res = await db.execute(
        select(MonthlyServiceLock).where(
            MonthlyServiceLock.user_id == user.id,
            MonthlyServiceLock.year == year,
            MonthlyServiceLock.month == month,
        )
    )
    lock = res.scalar_one_or_none()
    if lock:
        await db.delete(lock)
        await db.commit()

    # Інвалідуємо кеш аналітики для цього періоду
    for key in list(_analytics_cache.keys()):
        if key.startswith(f"{user.id}:{year}:{month}:"):
            _analytics_cache.pop(key, None)

    return {"is_locked": False}


# ── Експорт Excel ──────────────────────────────────────────────────


@router.post("/export")
async def export_excel(
    body: ExportRequest,
    db: AsyncSession = Depends(get_db),
    user: User = Depends(get_current_user),
):
    data = await _build_analytics(db, user.id, body.year, body.month, body.doctor_id)

    wb = openpyxl.Workbook()
    month_label = f"{MONTHS_UA[body.month]} {body.year}"

    # ── Аркуш 1: Summary ──
    ws1 = wb.active
    ws1.title = "Summary"
    ws1.append([f"Платні послуги · {month_label}"])
    ws1.append([])
    ws1.append(["Показник", "Значення (грн)"])
    d = data.dashboard
    rows_sum = [
        ("Оборот (сума послуг)", d.total_revenue),
        ("Кількість послуг", d.total_quantity),
        ("Середній чек", d.avg_check),
        ("Дохід лікаря/лікарів", d.doctor_income),
        ("Витрати на матеріали", d.materials_cost),
        (f"Єдиний податок ({data.ep_rate}%)", d.ep_amount),
        (f"Військовий збір ({data.vz_rate}%)", d.vz_amount),
        ("Сумарні витрати", d.total_costs),
        ("Дохід організації", d.org_income),
        ("Готівка в касі", d.cash_in_register),
        ("Кошти на рахунку в банку", d.bank_amount),
    ]
    for row in rows_sum:
        ws1.append(list(row))

    # ── Аркуш 2: Services ──
    ws2 = wb.create_sheet("Services")
    hdr = ["Код", "Назва", "Ціна", "К-ть", "Сума", "Матеріали",
           f"ЄП ({data.ep_rate}%)", f"ВЗ ({data.vz_rate}%)",
           "До розподілу", "Дохід лікаря", "Дохід орг."]
    ws2.append(hdr)
    for row in data.services_table:
        ws2.append([
            row.code, row.name, row.price, row.total_quantity,
            row.sum, row.materials, row.ep_amount, row.vz_amount,
            row.to_split, row.doctor_income, row.org_income,
        ])

    # ── Аркуш 3: Costs ──
    ws3 = wb.create_sheet("Costs")
    ws3.append(["ТОП матеріалів"])
    ws3.append(["Назва", "Одиниця", "К-ть", "Вартість (грн)", "Частка (%)"])
    for m in data.top_materials:
        ws3.append([m.name, m.unit, m.total_quantity, m.total_cost, m.share_pct])
    ws3.append([])
    ws3.append(["Податки"])
    ws3.append([f"Єдиний податок ({data.ep_rate}%)", d.ep_amount])
    ws3.append([f"Військовий збір ({data.vz_rate}%)", d.vz_amount])

    # ── Аркуш 4: Doctors (якщо всі лікарі) ──
    if not body.doctor_id and data.reports:
        ws4 = wb.create_sheet("Doctors")
        ws4.append(["Лікар", "Готівка в касі", "Статус"])
        for rep in data.reports:
            ws4.append([rep.doctor_name, rep.cash_in_register, rep.status])

    # ── Автоширина ──
    for ws in wb.worksheets:
        for col in ws.columns:
            max_len = max(len(str(cell.value or "")) for cell in col)
            ws.column_dimensions[col[0].column_letter].width = max(max_len + 2, 10)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    filename = f"paid_services_{body.year}_{body.month:02d}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


# ── Поширення (Share) ──────────────────────────────────────────────


@router.post("/share", response_model=ShareResponse)
async def create_share(
    body: ShareCreate,
    db: AsyncSession = Depends(get_db),
    user: User = Depends(get_current_user),
):
    # Видалити протерміновані записи (lazy cleanup)
    await _cleanup_expired(db, user.id)

    data = await _build_analytics(db, user.id, body.year, body.month, body.doctor_id)

    doctor_label = "Всі лікарі"
    if body.doctor_id:
        dr = await _get_doctor_name(db, body.doctor_id)
        doctor_label = dr

    token = secrets.token_urlsafe(32)
    expires_at = datetime.now(timezone.utc) + timedelta(days=30)

    share = ShareReport(
        user_id=user.id,
        token=token,
        filter_snapshot={
            "doctor_id": body.doctor_id,
            "doctor_label": doctor_label,
            "year": body.year,
            "month": body.month,
        },
        payload_snapshot=data.model_dump(),
        expires_at=expires_at,
        is_deleted=False,
    )
    db.add(share)
    await db.commit()

    return ShareResponse(
        token=token,
        url=f"/share/{token}",
        expires_at=expires_at,
    )


@router.get("/share/{token}/view", response_model=PublicShareData)
async def view_share(token: str, db: AsyncSession = Depends(get_db)):
    """Публічний ендпоінт (без авторизації) — перегляд share-сторінки."""
    r = await db.execute(
        select(ShareReport).where(
            ShareReport.token == token,
            ShareReport.is_deleted == False,
            ShareReport.expires_at > datetime.now(timezone.utc),
        )
    )
    share = r.scalar_one_or_none()
    if not share:
        raise HTTPException(404, detail="Посилання не знайдено або термін дії закінчився")

    fs = share.filter_snapshot
    month_name = MONTHS_UA[fs.get("month", 1)]
    year = fs.get("year", "")
    doctor_label = fs.get("doctor_label", "Всі лікарі")
    filter_label = f"{month_name} {year} · {doctor_label}"

    return PublicShareData(
        token=token,
        filter_label=filter_label,
        expires_at=share.expires_at,
        analytics=share.payload_snapshot,
    )


@router.delete("/share/{token}", status_code=204)
async def delete_share(
    token: str,
    db: AsyncSession = Depends(get_db),
    user: User = Depends(get_current_user),
):
    r = await db.execute(
        select(ShareReport).where(
            ShareReport.token == token,
            ShareReport.user_id == user.id,
        )
    )
    share = r.scalar_one_or_none()
    if not share:
        raise HTTPException(404, detail="Посилання не знайдено")
    share.is_deleted = True
    await db.commit()


@router.post("/cleanup-shares", status_code=204)
async def cleanup_shares(
    db: AsyncSession = Depends(get_db),
    user: User = Depends(get_current_user),
):
    """Видалити всі протерміновані share-записи поточного користувача."""
    await _cleanup_expired(db, user.id)


async def _cleanup_expired(db: AsyncSession, user_id: int) -> None:
    r = await db.execute(
        select(ShareReport).where(
            ShareReport.user_id == user_id,
            ShareReport.expires_at <= datetime.now(timezone.utc),
        )
    )
    for share in r.scalars().all():
        await db.delete(share)
    await db.commit()


# ── AI-аналіз зображень (платні послуги) ──────────────────────────────

_PS_SYSTEM = """\
Ти — спеціалізований OCR-асистент для обробки звітів про платні медичні послуги ФОП-лікаря.
Твоя задача — точно розпізнати дані зі скріншоту або фото звіту та повернути їх у форматі JSON.
Відповідай ТІЛЬКИ валідним JSON без markdown, коментарів чи пояснень."""

_PS_PROMPT = """\
<context>
Це скріншот або фото звіту про надані платні медичні послуги. Документ може містити:
- Таблицю з переліком послуг, кількістю та сумами
- Ім'я лікаря
- Суму готівки в касі
- Перелік наданих послуг із кількістю за період (місяць)
</context>

<task>
Знайди та вилучи з зображення:

1. **Ім'я лікаря** (якщо видно)
2. **Готівка в касі** — загальна сума готівки (число, грн)
3. **Перелік послуг** — для кожної послуги:
   - Назва послуги (точно як написано)
   - Кількість наданих послуг за період
</task>

<rules>
- Кількість — ціле число (integer)
- Готівка — число з можливими десятковими (float)
- Назви послуг — записуй ТОЧНО як на зображенні, не скорочуй
- Якщо дані не видно — пропускай послугу
- confidence: "high" = всі дані чітко видимі; "medium" = деякі дані неточні; "low" = більшість не розпізнано
</rules>

<output_format>
Поверни ТІЛЬКИ JSON:
{
  "doctor_name": "ПІБ або null",
  "cash_in_register": 0,
  "services": [
    {"name": "Назва послуги", "quantity": 0}
  ],
  "confidence": "high",
  "notes": ""
}
</output_format>"""


_PS_FALLBACK = {
    "doctor_name": None,
    "cash_in_register": 0,
    "services": [],
    "confidence": "low",
    "notes": "Не вдалося розпізнати дані з зображення",
}


@router.post("/analyze-image")
async def analyze_paid_services_image(
    images: list[UploadFile] = File(...),
    db: AsyncSession = Depends(get_db),
    user: User = Depends(get_current_user),
):
    """Аналізує зображення звіту платних послуг через AI."""
    provider, api_key = await get_provider(db, user.id)

    if not provider:
        raise HTTPException(
            status_code=503,
            detail="AI сервіс не налаштований. Додайте API ключ у Налаштуваннях.",
        )

    # Завантажуємо каталог послуг для fuzzy-matching
    svc_result = await db.execute(
        select(Service).where(Service.user_id == user.id).order_by(Service.code)
    )
    catalog = svc_result.scalars().all()
    catalog_names = {s.id: s.name.lower().strip() for s in catalog}

    results = []
    for image_file in images:
        raw = await image_file.read()
        media_type = image_file.content_type or "image/png"

        if len(raw) > 10 * 1024 * 1024:
            results.append({"error": "Файл перевищує 10 MB"})
            continue

        text = await ai_analyze_image(provider, api_key, raw, media_type, _PS_SYSTEM, _PS_PROMPT)
        data = parse_ai_json(text, _PS_FALLBACK)

        # Мепінг розпізнаних послуг на каталог (fuzzy)
        matched_entries: list[dict] = []
        ai_services = data.get("services", [])
        for ai_svc in ai_services:
            ai_name = (ai_svc.get("name") or "").lower().strip()
            qty = int(ai_svc.get("quantity", 0))
            if not ai_name or qty <= 0:
                continue
            best_id = None
            best_score = 0.0
            for sid, cname in catalog_names.items():
                # Simple substring/overlap matching
                if ai_name in cname or cname in ai_name:
                    score = len(cname) / max(len(ai_name), 1)
                    if score > best_score:
                        best_score = score
                        best_id = sid
                else:
                    # Word overlap
                    ai_words = set(ai_name.split())
                    c_words = set(cname.split())
                    overlap = len(ai_words & c_words)
                    total = max(len(ai_words | c_words), 1)
                    score = overlap / total
                    if score > best_score and score >= 0.4:
                        best_score = score
                        best_id = sid
            if best_id:
                matched_entries.append({"service_id": best_id, "quantity": qty})

        results.append({
            "doctor_name": data.get("doctor_name"),
            "cash_in_register": float(data.get("cash_in_register", 0)),
            "entries": matched_entries,
            "raw_services": ai_services,
            "confidence": data.get("confidence", "low"),
            "notes": data.get("notes", ""),
            "provider": provider,
        })

    return {"results": results, "provider": provider}
